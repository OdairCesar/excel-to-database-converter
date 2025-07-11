#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import pandas as pd
import sys
import os
import argparse
import platform
from datetime import datetime

try:
    from sql_schema_parser import SQLSchemaParser
except ImportError:
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    from sql_schema_parser import SQLSchemaParser

def show_banner():
    print("=" * 60)
    print("CONVERSOR EXCEL PARA BANCO DE DADOS UNIVERSAL")
    print("Suporte a schemas SQL externos")
    print("=" * 60)
    print()

def convert_excel_to_database(excel_file, output_dir="./conversao_output", schema_file=None, verbose=False, xlsx_only=False):
    """Converte Excel para banco de dados usando schema SQL"""
    
    # Carregar schema
    if schema_file:
        parser = SQLSchemaParser()
        parsed_schema = parser.parse_file(schema_file)
        schema = convert_schema_format(parsed_schema)
    else:
        schema = get_default_schema()
    
    if verbose:
        print(f"Schema: {schema['table_name']}")
        print(f"Campos: {list(schema['fields'].keys())}")
    
    # Ler Excel
    df = pd.read_excel(excel_file, engine='openpyxl')
    
    if verbose:
        print(f"Excel - Registros: {len(df)}")
        print(f"Excel - Colunas: {list(df.columns)}")
    
    # Mapear dados
    result = pd.DataFrame()
    mapped_count = 0
    
    for field_name in schema['fields'].keys():
        src_col = find_matching_column(df.columns, field_name)
        
        if src_col:
            field_type = schema['fields'][field_name]['type'].upper()
            
            # Tratar tipos especiais do MySQL primeiro
            if 'SET(' in field_type or 'ENUM(' in field_type:
                # Para campos SET/ENUM, manter como string limpa
                result[field_name] = clean_string_values(df[src_col])
            elif 'INT' in field_type and field_name != 'id':
                # Para campos INT, tentar converter com tratamento de erro
                try:
                    result[field_name] = pd.to_numeric(df[src_col], errors='coerce').fillna(0).astype(int)
                except (ValueError, TypeError):
                    # Se não conseguir converter, manter como string limpa
                    result[field_name] = clean_string_values(df[src_col])
            elif 'VARCHAR' in field_type or 'TEXT' in field_type or 'CHAR' in field_type:
                # Para campos de texto, usar limpeza especial e truncamento
                cleaned_values = clean_string_values(df[src_col])
                result[field_name] = cleaned_values.apply(lambda x: truncate_string_by_schema(x, field_name, schema))
            elif 'DECIMAL' in field_type or 'REAL' in field_type:
                result[field_name] = pd.to_numeric(df[src_col], errors='coerce').fillna(0.0)
            elif 'TINYINT' in field_type:
                try:
                    result[field_name] = pd.to_numeric(df[src_col], errors='coerce').fillna(1).astype(int)
                except (ValueError, TypeError):
                    result[field_name] = clean_string_values(df[src_col])
            elif 'BIGINT' in field_type:
                try:
                    result[field_name] = pd.to_numeric(df[src_col], errors='coerce').fillna(0).astype(int)
                except (ValueError, TypeError):
                    result[field_name] = clean_string_values(df[src_col])
            elif 'DATE' in field_type or 'DATETIME' in field_type or 'TIMESTAMP' in field_type:
                # Para campos de data, tentar converter ou manter como string
                try:
                    result[field_name] = pd.to_datetime(df[src_col], errors='coerce')
                except:
                    result[field_name] = clean_string_values(df[src_col])
            elif 'JSON' in field_type:
                # Para campos JSON, manter como string limpa
                result[field_name] = clean_string_values(df[src_col])
            else:
                # Para outros tipos, manter como está
                result[field_name] = df[src_col]
            
            mapped_count += 1
            if verbose:
                print(f"Mapeado: {field_name} <- {src_col} (tipo: {field_type})")
        else:
            # Valores padrão
            if field_name == 'id':
                result[field_name] = range(1, len(df) + 1)
            elif field_name == 'ativo':
                result[field_name] = 'S'  # Valor padrão para SET
            elif field_name == 'created' or field_name == 'created_at':
                result[field_name] = datetime.now()
            elif field_name == 'lastupdate':
                result[field_name] = datetime.now()
            else:
                result[field_name] = None
    
    if verbose:
        print(f"Campos mapeados: {mapped_count}/{len(schema['fields'])}")
    
    # Salvar arquivos
    filename = os.path.splitext(os.path.basename(excel_file))[0]
    os.makedirs(output_dir, exist_ok=True)
    
    # Sempre gerar XLSX corrigido
    xlsx_file = os.path.join(output_dir, f"{filename}_corrigido.xlsx")
    save_corrected_xlsx(result, xlsx_file, excel_file)
    
    if not xlsx_only:
        # CSV
        csv_file = os.path.join(output_dir, f"{filename}_convertido.csv")
        result.to_csv(csv_file, index=False, encoding='utf-8')
        
        # SQL
        sql_file = os.path.join(output_dir, f"{filename}_inserts.sql")
        generate_sql_file(result, schema, sql_file, excel_file)
        
        print(f"Arquivos gerados:")
        print(f"  CSV: {csv_file}")
        print(f"  SQL: {sql_file}")
        print(f"  XLSX: {xlsx_file}")
    else:
        print(f"Arquivo XLSX corrigido gerado:")
        print(f"  XLSX: {xlsx_file}")
    
    return result

def convert_schema_format(parsed_schema):
    """Converte schema do parser para formato usado pelo conversor"""
    converted_fields = {}
    
    for field_name, field_info in parsed_schema['fields'].items():
        field_type = field_info['type'].upper()
        
        if field_info.get('size'):
            field_type += f"({field_info['size']})"
        elif field_info.get('precision') and field_info.get('scale'):
            field_type += f"({field_info['precision']},{field_info['scale']})"
        
        constraints = []
        if not field_info.get('nullable', True):
            constraints.append('NOT NULL')
        if field_info.get('default'):
            constraints.append(f'DEFAULT {field_info["default"]}')
        if field_info.get('auto_increment'):
            constraints.append('AUTO_INCREMENT')
        
        converted_fields[field_name] = {
            'type': field_type,
            'constraints': constraints
        }
    
    return {
        'table_name': parsed_schema['table_name'],
        'fields': converted_fields,
        'primary_keys': parsed_schema.get('primary_keys', []),
        'indexes': parsed_schema.get('indexes', [])
    }

def get_default_schema():
    """Schema padrao cadastro"""
    return {
        'table_name': 'cadastro',
        'fields': {
            'id': {'type': 'INT', 'constraints': ['PRIMARY KEY', 'AUTO_INCREMENT']},
            'nome': {'type': 'VARCHAR(255)', 'constraints': ['NOT NULL']},
            'email': {'type': 'VARCHAR(100)', 'constraints': []},
            'ativo': {'type': 'TINYINT(1)', 'constraints': ['DEFAULT 1']},
            'created_at': {'type': 'TIMESTAMP', 'constraints': ['DEFAULT CURRENT_TIMESTAMP']}
        },
        'primary_keys': ['id'],
        'indexes': []
    }

def find_matching_column(excel_columns, target_field):
    """Encontra coluna correspondente no Excel"""
    # Correspondencia exata
    for col in excel_columns:
        if col.lower() == target_field.lower():
            return col
    
    # Correspondencia parcial
    field_variants = {
        'nome': ['nome', 'name', 'produto', 'item'],
        'preco': ['preco', 'price', 'valor', 'custo'],
        'categoria': ['categoria', 'category', 'tipo', 'class'],
        'descricao': ['descricao', 'description', 'desc', 'detalhes'],
        'email': ['email', 'e-mail', 'e_mail'],
        'ativo': ['ativo', 'active', 'status']
    }
    
    if target_field in field_variants:
        for variant in field_variants[target_field]:
            for col in excel_columns:
                if variant.lower() in col.lower():
                    return col
    
    return None

def generate_sql_file(dataframe, schema, sql_file, original_file):
    """Gera arquivo SQL"""
    table_name = schema['table_name']
    
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write(f"-- Conversao de {original_file}\n")
        f.write(f"-- Registros: {len(dataframe)}\n")
        f.write(f"-- Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"-- Tabela: {table_name}\n\n")
        
        # CREATE TABLE
        f.write(f"-- Estrutura da tabela {table_name}\n")
        f.write(generate_create_table_sql(schema))
        f.write("\n")
        
        # INSERTs
        f.write(f"-- Dados da tabela {table_name}\n")
        for index, row in dataframe.iterrows():
            columns = []
            values = []
            
            for field_name in schema['fields'].keys():
                if field_name in dataframe.columns:
                    columns.append(f"`{field_name}`")
                    value = row[field_name]
                    
                    if pd.isna(value):
                        values.append("NULL")
                    elif isinstance(value, (int, float)) and not pd.isna(value):
                        values.append(str(value))
                    elif isinstance(value, datetime):
                        values.append(f"'{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                    else:
                        escaped_value = str(value).replace("'", "''")
                        values.append(f"'{escaped_value}'")
            
            if columns:
                f.write(f"INSERT INTO `{table_name}` ({', '.join(columns)}) VALUES ({', '.join(values)});\n")
        
        f.write("\n-- Fim da conversao\n")

def generate_create_table_sql(schema):
    """Gera CREATE TABLE"""
    table_name = schema['table_name']
    fields = schema['fields']
    
    sql = f"CREATE TABLE IF NOT EXISTS `{table_name}` (\n"
    field_definitions = []
    
    for field_name, field_info in fields.items():
        field_type = field_info['type']
        constraints = field_info.get('constraints', [])
        
        definition = f"  `{field_name}` {field_type}"
        if constraints:
            definition += f" {' '.join(constraints)}"
        
        field_definitions.append(definition)
    
    sql += ",\n".join(field_definitions)
    sql += "\n);\n"
    
    return sql

def clean_string_values(series):
    """Limpa valores de string removendo .0 indesejados e NaN"""
    # Converter para string
    cleaned = series.astype(str)
    # Remover 'nan' e 'None'
    cleaned = cleaned.replace(['nan', 'None'], '')
    # Remover .0 no final de números que deveriam ser strings
    cleaned = cleaned.str.replace(r'\.0$', '', regex=True)
    # Remover espaços extras
    cleaned = cleaned.str.strip()
    return cleaned

def truncate_string_by_schema(value, field_name, schema):
    """Trunca string baseada no tamanho definido no schema"""
    if pd.isna(value) or value == '':
        return value
    
    value_str = str(value)
    
    # Verificar se o campo existe no schema
    if field_name not in schema['fields']:
        return value_str
    
    field_type = schema['fields'][field_name]['type'].upper()
    
    # Extrair tamanho do campo VARCHAR, CHAR, etc.
    import re
    size_match = re.search(r'\((\d+)\)', field_type)
    if size_match:
        max_size = int(size_match.group(1))
        if len(value_str) > max_size:
            truncated = value_str[:max_size]
            print(f"⚠️  Campo '{field_name}' truncado: '{value_str}' → '{truncated}' (limite: {max_size})")
            return truncated
    
    return value_str

def save_corrected_xlsx(dataframe, output_file, original_file):
    """Salva DataFrame corrigido em arquivo XLSX"""
    try:
        # Usar openpyxl para melhor controle
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Corrigidos"
        
        # Adicionar cabeçalho com informações
        ws['A1'] = f"Arquivo Original: {os.path.basename(original_file)}"
        ws['A2'] = f"Registros: {len(dataframe)}"
        ws['A3'] = f"Gerado em: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = ""  # Linha vazia
        
        # Estilo para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Adicionar dados a partir da linha 5
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 5):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Aplicar estilo ao cabeçalho
                if r_idx == 5:  # Linha de cabeçalho dos dados
                    cell.font = header_font
                    cell.fill = header_fill
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar arquivo
        wb.save(output_file)
        print(f"✅ Arquivo XLSX corrigido salvo: {output_file}")
        
    except ImportError:
        # Fallback para pandas simples
        print("⚠️  Usando pandas para salvar XLSX (openpyxl não disponível)")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name='Dados Corrigidos', index=False)
        print(f"✅ Arquivo XLSX corrigido salvo: {output_file}")
    
    except Exception as e:
        print(f"❌ Erro ao salvar XLSX: {e}")
        # Tentar salvar com pandas simples
        try:
            dataframe.to_excel(output_file, index=False)
            print(f"✅ Arquivo XLSX corrigido salvo (modo simples): {output_file}")
        except Exception as e2:
            print(f"❌ Erro crítico ao salvar XLSX: {e2}")
            raise

def test_conversion():
    print("Testando conversao...")
    
    # Testar parser SQL
    parser = SQLSchemaParser()
    schema = parser.parse_file('schemas/produtos.sql')
    
    print(f"Schema: {schema['table_name']}")
    print(f"Fields: {list(schema['fields'].keys())}")
    
    # Testar leitura do Excel
    df = pd.read_excel('test_produtos.xlsx')
    print(f"Excel columns: {list(df.columns)}")
    print(f"Records: {len(df)}")
    
    return True

def main():
    parser = argparse.ArgumentParser(
        description="Conversor Excel para banco de dados usando schemas SQL",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos de uso:
  python convert_excel_universal.py arquivo.xlsx
  python convert_excel_universal.py arquivo.xlsx --schema schemas/produtos.sql
  python convert_excel_universal.py arquivo.xlsx -o ./saida --verbose
        """
    )
    parser.add_argument('excel_file', help='Arquivo Excel')
    parser.add_argument('--schema', help='Arquivo SQL schema')
    parser.add_argument('-o', '--output', default='./conversao_output', help='Diretorio de saida')
    parser.add_argument('-v', '--verbose', action='store_true', help='Verbose')
    
    args = parser.parse_args()
    
    show_banner()
    
    print(f"Sistema: {platform.system()}")
    print(f"Arquivo: {args.excel_file}")
    if args.schema:
        print(f"Schema: {args.schema}")
    print(f"Saida: {args.output}")
    print()
    
    print("Executando conversao...")
    try:
        result = convert_excel_to_database(args.excel_file, args.output, args.schema, args.verbose)
        print(f"? Conversao concluida! ({len(result)} registros)")
        return 0
    except Exception as e:
        print(f"? Erro na conversao: {e}")
        return 1

if __name__ == "__main__":
    sys.exit(main())
